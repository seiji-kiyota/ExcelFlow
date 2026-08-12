"""Export processed data to Excel workbooks."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, MutableMapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_flow.validators import (
    ExcelFlowError,
    validate_export_filename,
    validate_export_selection,
)

SHEET_DATA = "整形済データ"
SHEET_AGGREGATED = "集計結果"
SHEET_HISTORY = "処理履歴"

# Widget key for the filename text input (must not collide with download metadata keys).
EXPORT_FILENAME_KEY = "export_filename_input"
EXPORT_INCLUDE_DATA_KEY = "export_include_data"
EXPORT_INCLUDE_AGG_KEY = "export_include_aggregated"
EXPORT_INCLUDE_HISTORY_KEY = "export_include_history"

# Download-ready artifacts (separate from widget keys).
EXPORT_BYTES_KEY = "export_bytes"
EXPORT_READY_FILENAME_KEY = "export_ready_filename"
EXPORT_CONFIG_KEY = "export_config"
EXPORT_READY_KEY = "export_ready"

HEADER_FILL = PatternFill("solid", fgColor="1C83E1")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MAX_COLUMN_WIDTH = 40
MIN_COLUMN_WIDTH = 10


def clear_export_artifacts(session_state: MutableMapping) -> None:
    """Clear generated Excel artifacts only (keep Phase 6 widget selections)."""
    for key in (
        EXPORT_BYTES_KEY,
        EXPORT_READY_FILENAME_KEY,
        EXPORT_CONFIG_KEY,
        EXPORT_READY_KEY,
        # Legacy key used before the widget-key collision fix.
        "export_filename",
    ):
        session_state.pop(key, None)


def store_export_result(
    session_state: MutableMapping,
    *,
    export_bytes: bytes,
    filename: str,
    export_config: dict[str, Any],
) -> None:
    """Store a successfully generated Excel file for download."""
    session_state[EXPORT_BYTES_KEY] = export_bytes
    session_state[EXPORT_READY_FILENAME_KEY] = filename
    session_state[EXPORT_CONFIG_KEY] = export_config
    session_state[EXPORT_READY_KEY] = True


def is_export_ready(session_state: MutableMapping) -> bool:
    """Return True when a downloadable Excel result is available."""
    return bool(
        session_state.get(EXPORT_READY_KEY)
        and session_state.get(EXPORT_BYTES_KEY)
        and session_state.get(EXPORT_READY_FILENAME_KEY)
    )


def reset_export_state(session_state: MutableMapping) -> None:
    """Clear Phase 6 export artifacts and widget keys.

    Call before Phase 6 widgets are instantiated when resetting via callback.
    """
    clear_export_artifacts(session_state)
    for key in (
        EXPORT_FILENAME_KEY,
        EXPORT_INCLUDE_DATA_KEY,
        EXPORT_INCLUDE_AGG_KEY,
        EXPORT_INCLUDE_HISTORY_KEY,
    ):
        session_state.pop(key, None)


def generate_default_filename(now: datetime | None = None) -> str:
    """Return ``ExcelFlow_Result_YYYYMMDD_HHMMSS.xlsx``."""
    stamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return f"ExcelFlow_Result_{stamp}.xlsx"


def normalize_export_filename(filename: str | None) -> str:
    """Validate and normalize a user-provided export filename."""
    return validate_export_filename(filename)


def build_process_history(
    metadata: dict[str, Any],
) -> pd.DataFrame:
    """Build a two-column process-history table from metadata."""
    rows = metadata.get("rows")
    if rows is None:
        # Accept ordered list of (item, value) or dict items.
        items = metadata.get("items")
        if isinstance(items, dict):
            rows = [{"項目": key, "内容": value} for key, value in items.items()]
        elif isinstance(items, list):
            rows = items
        else:
            rows = []
    history = pd.DataFrame(rows, columns=["項目", "内容"])
    return history


def build_process_history_from_session(context: dict[str, Any]) -> pd.DataFrame:
    """Create a readable process history from app/session context."""
    file_meta = context.get("file_meta") or {}
    cleaning_summary = context.get("cleaning_summary")
    cleaning_config = context.get("cleaning_config") or {}
    aggregation_config = context.get("aggregation_config")
    chart_config = context.get("chart_config")
    chart_generated = bool(context.get("chart_generated"))
    data_source_label = context.get("data_source_label", "元データ")
    original_df = context.get("original_df")
    cleaned_df = context.get("cleaned_df")
    aggregated_df = context.get("aggregated_df")
    sort_label = context.get("sort_label", "未実施")

    dropped = cleaning_config.get("columns_to_drop") or []
    renames = cleaning_config.get("rename_mapping") or {}
    rename_text = ", ".join(f"{src} → {dst}" for src, dst in renames.items()) or "未実施"
    dropped_text = ", ".join(str(col) for col in dropped) or "未実施"

    has_cleaning = cleaned_df is not None and cleaning_summary is not None
    has_aggregation = aggregated_df is not None and aggregation_config is not None

    rows = [
        {"項目": "元ファイル名", "内容": file_meta.get("filename", "未設定")},
        {"項目": "ファイル形式", "内容": str(file_meta.get("file_format", "未設定")).upper()},
        {"項目": "Excelシート名", "内容": file_meta.get("sheet_name") or "（CSVのためなし）"},
        {
            "項目": "読込行数",
            "内容": len(original_df) if original_df is not None else "未設定",
        },
        {
            "項目": "読込列数",
            "内容": len(original_df.columns) if original_df is not None else "未設定",
        },
        {"項目": "出力対象データ", "内容": data_source_label},
        {"項目": "整形実行", "内容": "実施" if has_cleaning else "未実施"},
        {"項目": "削除列", "内容": dropped_text if has_cleaning else "未実施"},
        {"項目": "列名変更", "内容": rename_text if has_cleaning else "未実施"},
        {
            "項目": "前後空白除去",
            "内容": (
                "ON"
                if has_cleaning and cleaning_config.get("strip_whitespace")
                else ("OFF" if has_cleaning else "未実施")
            ),
        },
        {
            "項目": "空白行削除",
            "内容": (
                "ON"
                if has_cleaning and cleaning_config.get("remove_blank_rows")
                else ("OFF" if has_cleaning else "未実施")
            ),
        },
        {
            "項目": "重複行削除",
            "内容": (
                "ON"
                if has_cleaning and cleaning_config.get("remove_duplicates")
                else ("OFF" if has_cleaning else "未実施")
            ),
        },
        {
            "項目": "整形前行数",
            "内容": cleaning_summary["original_rows"] if has_cleaning else "未実施",
        },
        {
            "項目": "整形後行数",
            "内容": cleaning_summary["cleaned_rows"] if has_cleaning else "未実施",
        },
        {
            "項目": "整形前列数",
            "内容": cleaning_summary["original_columns"] if has_cleaning else "未実施",
        },
        {
            "項目": "整形後列数",
            "内容": cleaning_summary["cleaned_columns"] if has_cleaning else "未実施",
        },
        {"項目": "集計実行", "内容": "実施" if has_aggregation else "未実施"},
        {
            "項目": "グループ項目",
            "内容": (
                " × ".join(aggregation_config["group_columns"])
                if has_aggregation
                else "未実施"
            ),
        },
        {
            "項目": "集計方法",
            "内容": (
                aggregation_config.get("aggregation_label", "未実施")
                if has_aggregation
                else "未実施"
            ),
        },
        {
            "項目": "集計対象列",
            "内容": (
                (
                    aggregation_config.get("value_column")
                    or aggregation_config.get("result_column")
                    or "未実施"
                )
                if has_aggregation
                else "未実施"
            ),
        },
        {"項目": "並べ替え", "内容": sort_label if has_aggregation else "未実施"},
        {
            "項目": "集計結果件数",
            "内容": len(aggregated_df) if has_aggregation else "未実施",
        },
        {"項目": "グラフ作成", "内容": "実施" if chart_generated and chart_config else "未実施"},
        {
            "項目": "グラフ種類",
            "内容": (
                {
                    "bar": "棒グラフ",
                    "line": "折れ線グラフ",
                    "pie": "円グラフ",
                }.get(str(chart_config.get("chart_type")), chart_config.get("chart_type"))
                if chart_generated and chart_config
                else "未実施"
            ),
        },
        {
            "項目": "グラフタイトル",
            "内容": (
                chart_config.get("title") or "（なし）"
                if chart_generated and chart_config
                else "未実施"
            ),
        },
        {"項目": "出力日時", "内容": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]
    return pd.DataFrame(rows)


def _style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _autosize_columns(worksheet, dataframe: pd.DataFrame) -> None:
    for index, column_name in enumerate(dataframe.columns, start=1):
        series = dataframe.iloc[:, index - 1].head(200)
        lengths = [len(str(column_name))]
        for value in series.tolist():
            if value is None or (isinstance(value, float) and pd.isna(value)):
                continue
            try:
                if pd.isna(value):
                    continue
            except (TypeError, ValueError):
                pass
            lengths.append(len(str(value)))
        max_len = max(lengths, default=MIN_COLUMN_WIDTH)
        width = min(max(max_len + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _apply_number_and_date_formats(worksheet, dataframe: pd.DataFrame) -> None:
    for col_index, column_name in enumerate(dataframe.columns, start=1):
        series = dataframe[column_name]
        if pd.api.types.is_datetime64_any_dtype(series):
            number_format = "yyyy-mm-dd"
        elif pd.api.types.is_integer_dtype(series):
            number_format = "#,##0"
        elif pd.api.types.is_float_dtype(series):
            number_format = "#,##0.00"
        else:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_index, column=col_index).number_format = number_format


def _finalize_data_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    _style_header(worksheet)
    _autosize_columns(worksheet, dataframe)
    _apply_number_and_date_formats(worksheet, dataframe)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _finalize_history_sheet(worksheet, dataframe: pd.DataFrame) -> None:
    _style_header(worksheet)
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 50
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _ordered_sheet_names(
    *,
    include_aggregated: bool,
    include_data: bool,
    include_history: bool,
) -> list[str]:
    """Return the canonical worksheet tab order for export."""
    ordered: list[str] = []
    if include_aggregated:
        ordered.append(SHEET_AGGREGATED)
    if include_data:
        ordered.append(SHEET_DATA)
    if include_history:
        ordered.append(SHEET_HISTORY)
    return ordered


def _apply_sheet_order_and_active(workbook, ordered_names: list[str]) -> None:
    """Force worksheet tab order and active sheet before the file is saved."""
    if not ordered_names:
        return
    sheets_by_title = {worksheet.title: worksheet for worksheet in workbook.worksheets}
    missing = [name for name in ordered_names if name not in sheets_by_title]
    if missing:
        raise ExcelFlowError(
            "Excelファイルを作成できませんでした。",
            detail=f"missing sheets for reorder: {missing}",
        )
    # openpyxl stores tab order in the internal _sheets list.
    workbook._sheets = [sheets_by_title[name] for name in ordered_names]
    workbook.active = workbook[ordered_names[0]]


def build_excel_workbook(
    data_df: pd.DataFrame | None,
    *,
    aggregated_df: pd.DataFrame | None = None,
    process_history: pd.DataFrame | None = None,
    include_data: bool = True,
    include_aggregated: bool = True,
    include_history: bool = True,
) -> BytesIO:
    """Build an Excel workbook in memory and return a BytesIO buffer."""
    include_data, include_aggregated, include_history = validate_export_selection(
        data_df=data_df,
        aggregated_df=aggregated_df,
        include_data=include_data,
        include_aggregated=include_aggregated,
        include_history=include_history,
    )

    buffer = BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Write in the same order as the final tab order.
            if include_aggregated:
                assert aggregated_df is not None
                agg_copy = aggregated_df.copy()
                agg_copy.to_excel(writer, sheet_name=SHEET_AGGREGATED, index=False)
                _finalize_data_sheet(writer.sheets[SHEET_AGGREGATED], agg_copy)

            if include_data:
                assert data_df is not None
                data_copy = data_df.copy()
                data_copy.to_excel(writer, sheet_name=SHEET_DATA, index=False)
                _finalize_data_sheet(writer.sheets[SHEET_DATA], data_copy)

            if include_history:
                history = (
                    process_history.copy()
                    if process_history is not None
                    else pd.DataFrame([{"項目": "処理履歴", "内容": "記録なし"}])
                )
                if list(history.columns) != ["項目", "内容"]:
                    history = build_process_history({"rows": history.to_dict(orient="records")})
                history.to_excel(writer, sheet_name=SHEET_HISTORY, index=False)
                _finalize_history_sheet(writer.sheets[SHEET_HISTORY], history)

            ordered_names = _ordered_sheet_names(
                include_aggregated=include_aggregated,
                include_data=include_data,
                include_history=include_history,
            )
            _apply_sheet_order_and_active(writer.book, ordered_names)
    except ExcelFlowError:
        raise
    except Exception as exc:  # noqa: BLE001 - convert to user-facing error
        raise ExcelFlowError(
            "Excelファイルを作成できませんでした。",
            detail=str(exc),
        ) from exc

    buffer.seek(0)
    return buffer
