"""Tests for excel_flow.excel_exporter."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from excel_flow.excel_exporter import (
    EXPORT_BYTES_KEY,
    EXPORT_FILENAME_KEY,
    EXPORT_READY_FILENAME_KEY,
    EXPORT_READY_KEY,
    SHEET_AGGREGATED,
    SHEET_DATA,
    SHEET_HISTORY,
    build_excel_workbook,
    build_process_history_from_session,
    clear_export_artifacts,
    generate_default_filename,
    is_export_ready,
    normalize_export_filename,
    reset_export_state,
    store_export_result,
)
from excel_flow.validators import ExcelFlowError


def _sample_data() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "部署": ["営業1課", "営業2課"],
            "数量": [2, 5],
            "売上": [100, 250],
        }
    )


def _sample_aggregated() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "部署": ["営業2課", "営業1課"],
            "件数": [3, 2],
        }
    )


def test_generate_default_filename() -> None:
    name = generate_default_filename(datetime(2026, 8, 13, 7, 15, 0))
    assert name == "ExcelFlow_Result_20260813_071500.xlsx"


def test_normalize_export_filename_adds_extension() -> None:
    assert normalize_export_filename("result") == "result.xlsx"
    assert normalize_export_filename("result.xlsx") == "result.xlsx"


def test_normalize_export_filename_rejects_invalid_chars() -> None:
    with pytest.raises(ExcelFlowError) as exc_info:
        normalize_export_filename("bad:name")
    assert "使用できない文字" in exc_info.value.user_message


def test_normalize_export_filename_rejects_empty() -> None:
    with pytest.raises(ExcelFlowError):
        normalize_export_filename("   ")


def test_build_excel_workbook_all_sheets() -> None:
    data = _sample_data()
    aggregated = _sample_aggregated()
    history = pd.DataFrame([{"項目": "読込行数", "内容": 2}])
    snapshot_data = data.copy()
    snapshot_agg = aggregated.copy()

    buffer = build_excel_workbook(
        data,
        aggregated_df=aggregated,
        process_history=history,
        include_data=True,
        include_aggregated=True,
        include_history=True,
    )

    assert isinstance(buffer, BytesIO)
    pd.testing.assert_frame_equal(data, snapshot_data)
    pd.testing.assert_frame_equal(aggregated, snapshot_agg)

    workbook = load_workbook(buffer)
    assert workbook.sheetnames == [SHEET_AGGREGATED, SHEET_DATA, SHEET_HISTORY]
    assert workbook.active.title == SHEET_AGGREGATED

    data_ws = workbook[SHEET_DATA]
    assert data_ws.freeze_panes == "A2"
    assert data_ws.auto_filter.ref
    assert data_ws["A1"].font.bold
    assert data_ws.column_dimensions["A"].width >= 10
    assert data_ws["C2"].value == 100
    assert isinstance(data_ws["C2"].value, int)
    assert data_ws["C2"].number_format == "#,##0"

    agg_ws = workbook[SHEET_AGGREGATED]
    assert agg_ws["A2"].value == "営業2課"
    assert agg_ws["B2"].value == 3

    history_ws = workbook[SHEET_HISTORY]
    assert history_ws["A1"].value == "項目"
    assert history_ws["B2"].value == 2


def test_build_excel_without_aggregated() -> None:
    buffer = build_excel_workbook(
        _sample_data(),
        include_data=True,
        include_aggregated=False,
        include_history=True,
    )
    workbook = load_workbook(buffer)
    assert workbook.sheetnames == [SHEET_DATA, SHEET_HISTORY]
    assert workbook.active.title == SHEET_DATA
    assert SHEET_AGGREGATED not in workbook.sheetnames


def test_sheet_order_persists_after_disk_roundtrip(tmp_path) -> None:
    with_agg = build_excel_workbook(
        _sample_data(),
        aggregated_df=_sample_aggregated(),
        process_history=pd.DataFrame([{"項目": "読込行数", "内容": 2}]),
        include_data=True,
        include_aggregated=True,
        include_history=True,
    )
    path_with_agg = tmp_path / "with_agg.xlsx"
    path_with_agg.write_bytes(with_agg.getvalue())
    workbook = load_workbook(path_with_agg)
    assert workbook.sheetnames == [SHEET_AGGREGATED, SHEET_DATA, SHEET_HISTORY]
    assert workbook.active.title == SHEET_AGGREGATED

    without_agg = build_excel_workbook(
        _sample_data(),
        process_history=pd.DataFrame([{"項目": "読込行数", "内容": 2}]),
        include_data=True,
        include_aggregated=False,
        include_history=True,
    )
    path_without_agg = tmp_path / "without_agg.xlsx"
    path_without_agg.write_bytes(without_agg.getvalue())
    workbook2 = load_workbook(path_without_agg)
    assert workbook2.sheetnames == [SHEET_DATA, SHEET_HISTORY]
    assert workbook2.active.title == SHEET_DATA


def test_build_excel_without_history_is_safe() -> None:
    buffer = build_excel_workbook(
        _sample_data(),
        include_data=True,
        include_aggregated=False,
        include_history=False,
    )
    workbook = load_workbook(buffer)
    assert workbook.sheetnames == [SHEET_DATA]


def test_export_selection_requires_at_least_one_sheet() -> None:
    with pytest.raises(ExcelFlowError) as exc_info:
        build_excel_workbook(
            _sample_data(),
            include_data=False,
            include_aggregated=False,
            include_history=False,
        )
    assert "1つ以上選択" in exc_info.value.user_message


def test_export_rejects_missing_aggregated_when_requested() -> None:
    with pytest.raises(ExcelFlowError) as exc_info:
        build_excel_workbook(
            _sample_data(),
            aggregated_df=None,
            include_data=False,
            include_aggregated=True,
            include_history=False,
        )
    assert "集計結果がありません" in exc_info.value.user_message


def test_export_rejects_empty_dataframe() -> None:
    with pytest.raises(ExcelFlowError):
        build_excel_workbook(
            pd.DataFrame(),
            include_data=True,
            include_aggregated=False,
            include_history=False,
        )


def test_process_history_from_session() -> None:
    history = build_process_history_from_session(
        {
            "file_meta": {"filename": "sample_sales.xlsx", "file_format": "xlsx", "sheet_name": "売上データ"},
            "original_df": _sample_data(),
            "cleaned_df": _sample_data(),
            "cleaning_summary": {
                "original_rows": 2,
                "cleaned_rows": 2,
                "original_columns": 3,
                "cleaned_columns": 3,
            },
            "cleaning_config": {
                "columns_to_drop": ["備考"],
                "rename_mapping": {"担当者": "営業担当"},
                "strip_whitespace": True,
                "remove_blank_rows": False,
                "remove_duplicates": True,
            },
            "aggregated_df": _sample_aggregated(),
            "aggregation_config": {
                "group_columns": ["部署", "商品"],
                "aggregation_label": "件数",
                "value_column": None,
                "result_column": "件数",
            },
            "chart_generated": True,
            "chart_config": {"chart_type": "bar", "title": "部署別"},
            "data_source_label": "整形済データ",
            "sort_label": "降順",
        }
    )
    assert list(history.columns) == ["項目", "内容"]
    assert "元ファイル名" in set(history["項目"])
    assert "出力日時" in set(history["項目"])


def test_reset_export_state() -> None:
    state = {
        EXPORT_BYTES_KEY: b"abc",
        EXPORT_READY_FILENAME_KEY: "a.xlsx",
        "export_config": {"include_data": True},
        EXPORT_READY_KEY: True,
        "aggregated_df": _sample_aggregated(),
        "export_include_data": True,
        EXPORT_FILENAME_KEY: "a.xlsx",
    }
    reset_export_state(state)
    assert EXPORT_BYTES_KEY not in state
    assert EXPORT_READY_FILENAME_KEY not in state
    assert EXPORT_FILENAME_KEY not in state
    assert "aggregated_df" in state


def test_store_export_result_marks_ready() -> None:
    state: dict = {}
    store_export_result(
        state,
        export_bytes=b"xlsx-bytes",
        filename="result.xlsx",
        export_config={"include_data": True},
    )
    assert is_export_ready(state)
    assert state[EXPORT_BYTES_KEY] == b"xlsx-bytes"
    assert state[EXPORT_READY_FILENAME_KEY] == "result.xlsx"


def test_failed_export_does_not_keep_artifacts() -> None:
    state = {
        EXPORT_BYTES_KEY: b"old",
        EXPORT_READY_FILENAME_KEY: "old.xlsx",
        EXPORT_READY_KEY: True,
        "aggregated_df": _sample_aggregated(),
    }
    clear_export_artifacts(state)
    assert not is_export_ready(state)
    assert EXPORT_BYTES_KEY not in state
    assert "aggregated_df" in state


def test_failed_recreate_clears_previous_ready_export() -> None:
    """Simulate: old export exists, new create fails after clearing artifacts."""
    state = {}
    store_export_result(
        state,
        export_bytes=b"previous",
        filename="previous.xlsx",
        export_config={"include_data": True},
    )
    assert is_export_ready(state)

    # Creation start: clear first so failure cannot expose stale download.
    clear_export_artifacts(state)
    assert not is_export_ready(state)

    try:
        raise ExcelFlowError("Excelファイルを作成できませんでした。")
    except ExcelFlowError:
        clear_export_artifacts(state)

    assert not is_export_ready(state)
    assert EXPORT_BYTES_KEY not in state
